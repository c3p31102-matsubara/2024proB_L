import openpyxl
import os
import MySQLdb

path = os.path.dirname(__file__)
wb = openpyxl.load_workbook(os.path.join(path, "sheet.xlsx")) 

# データベースに接続
with MySQLdb.connect(host="localhost", user="probc2024", passwd="probc2024", db="probc2024") as con:
    with con.cursor() as cur:
        sql = """SELECT discovery.category, discovery.color, discovery.features, discovery.datetime, discovery.place
                from discovery"""
        cur.execute(sql)
        rows = cur.fetchall()
        i, p = 0, 0
        for row in rows:
            if (i + 10) > 46 or i == 0:
                i, p = 0, p + 1
                ws = wb.copy_worksheet(wb["template"])
                ws.title = f"拾得物リスト{p}"
            ws.cell(13+i, 9).value = row[0]
        else:
                ws.cell(13+i, 11).value = row[1] + row[2]
                ws.cell(13+i, 14).value = row[3].strftime('%m/%d %H:%M')
                ws.cell(13+i, 15).value = row[4]
                i += 4
        wb.remove(wb["template"]) 

wb.save(os.path.join(path, "out.xlsx"))  